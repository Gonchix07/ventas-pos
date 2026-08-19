"""
Importación del padrón de clientes del ERP anterior (Excel) a POS-Ventas.

Carga de una sola vez: lee el Excel, normaliza, resuelve los catálogos por descripción y hace la
inserción masiva de Clientes + TarjetasClientes en una sola transacción.

    python importar_padron_clientes.py --excel "D:\\Escritorio\\Clientes.xlsx" --dry-run
    python importar_padron_clientes.py --excel "D:\\Escritorio\\Clientes.xlsx"

La cadena de conexión sale de la variable de entorno POS_CONN (mismo formato que
ConnectionStrings__Pos) o de --conn. Con --dry-run no se toca la base: se transforma todo, se emite
el reporte de anomalías y se corta.

Decisiones tomadas con el usuario (11/08/2026):
  - Los 5 clientes preexistentes NO se borran: tienen ventas, convenios y cuentas corrientes
    asociadas, y ya están inactivos. El padrón se agrega al lado, con IdCliente desde MAX+1.
  - Las filas sin tipo de tarjeta (vacío o '00') se asignan a TARJETA ROJA.
  - Se crean las condiciones de IVA que faltan en el catálogo en vez de forzarlas a Consumidor Final.
"""
import argparse
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict

import openpyxl
import pyodbc

# --- Mapeo de condición de IVA del Excel a la descripción del catálogo -----------------------
# Se resuelve por descripción y no por id fijo: los ids del catálogo dependen del seed.
COND_IVA = {
    "CONSUMIDOR FINAL": "Consumidor Final",
    "REGIMEN SIMPLIFICADO": "Monotributista",          # régimen simplificado = monotributo
    "RESPONSABLE INSCRIPTO": "Responsable Inscripto",
    "EXENTO/NO ALCANZADO": "Exento",
    "RESP. INS M": "Responsable Inscripto",            # variante tipeada, 1 sola fila
    "RESPONSABLE NO INSCRIPTO": "Responsable No Inscripto",
    "SUJETO NO CATEGORIZADO": "Sujeto No Categorizado",
}
# Condiciones a crear si no existen. La letra define el comprobante a emitir: se usa 'B' (sin
# discriminar IVA, igual que Consumidor Final) porque son categorías derogadas de AFIP y las 5 filas
# que las traen son datos heredados. CONFIRMAR con el usuario antes de facturarles.
COND_IVA_NUEVAS = {
    "Responsable No Inscripto": ("B", "RNI"),
    "Sujeto No Categorizado": ("B", "SNC"),
}

TIPO_TARJETA = {"ROJA": "TARJETA ROJA", "AZUL": "TARJETA AZUL"}
TIPO_TARJETA_DEFECTO = "TARJETA ROJA"   # filas con tipo vacío o '00'

# Largos de las columnas destino (ver PosDbContext).
LARGOS = {
    "CodigoInt": 450, "Cuit": 11, "Documento": 450, "Descripcion": 200,
    "Domicilio": 120, "CodigoPostal": 8, "Localidad": 60, "Email": 120,
    "NroTarjeta": 450,
}

COLUMNAS = ["Codigo Cliente", "Razon Social", "Domicilio", "copostal", "localidad",
            "email", "cuit", "numdoc", "Condicion IVA", "Admite Presupuesto",
            "Estado", "tipo_tarjeta", "Codigo Tarjeta"]


class Reporte:
    """Junta las anomalías para volcarlas a CSV: nada se descarta en silencio."""

    def __init__(self):
        self.filas = []
        self.conteo = Counter()

    def add(self, tipo, codigo, detalle):
        self.conteo[tipo] += 1
        if self.conteo[tipo] <= 200:      # se guardan ejemplos, no las 150k
            self.filas.append((tipo, codigo, detalle))

    def volcar(self, ruta):
        import csv
        with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["tipo", "codigo_cliente", "detalle"])
            w.writerows(self.filas)

    def resumen(self):
        return self.conteo.most_common()


def reponer_enies(s):
    """
    El export del ERP anterior trae '±' donde iba una eñe ("SAENZ PE±A", "Saenz Pe±a"). Se repone
    respetando la caja del contexto: el padrón está casi todo en mayúsculas, pero hay ~670 domicilios
    con mayúscula inicial donde una 'Ñ' quedaría mal escrita dentro de una palabra en minúscula
    ("CataluÑa" en lugar de "Cataluña").
    """
    if "±" not in s:
        return s
    out = list(s)
    for i, ch in enumerate(out):
        if ch != "±":
            continue
        prev = s[i - 1] if i else ""
        nxt = s[i + 1] if i + 1 < len(s) else ""
        en_minuscula = (prev.isalpha() and prev.islower()) or (nxt.isalpha() and nxt.islower())
        out[i] = "ñ" if en_minuscula else "Ñ"
    return "".join(out)


def limpiar_texto(v):
    """Normaliza un valor de celda a texto: sin espacios sobrantes ni caracteres de control."""
    if v is None:
        return ""
    s = str(v).strip()
    # Los demás acentuados del origen (Ñ, º, °, Ó, Í) son legítimos y se conservan; solo se descarta
    # el carácter de reemplazo suelto.
    s = reponer_enies(s).replace("�", "")
    s = "".join(ch for ch in s if unicodedata.category(ch)[0] != "C")
    return re.sub(r"\s{2,}", " ", s).strip()


def recortar(valor, campo, codigo, rep):
    """Recorta al largo de la columna destino, dejando constancia."""
    lim = LARGOS[campo]
    if len(valor) > lim:
        rep.add(f"truncado:{campo}", codigo, f"{len(valor)} caracteres -> {lim}: {valor!r}")
        return valor[:lim]
    return valor


def transformar(ruta_excel, rep):
    """Excel -> (clientes, tarjetas). Sin tocar la base."""
    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
    if "clientes" not in wb.sheetnames:
        raise SystemExit(f"El Excel no tiene la hoja 'clientes' (tiene {wb.sheetnames})")
    ws = wb["clientes"]
    it = ws.iter_rows(min_row=1, values_only=True)
    headers = [limpiar_texto(h) for h in next(it)]
    if headers[:len(COLUMNAS)] != COLUMNAS:
        raise SystemExit(f"Encabezados inesperados.\n  esperado: {COLUMNAS}\n  recibido: {headers}")
    H = {h: i for i, h in enumerate(headers)}

    clientes = {}                       # CodigoInt -> dict
    tarjetas = set()                    # (CodigoInt, tipoDescripcion, NroTarjeta)
    vistos_cond = Counter()

    for nro, fila in enumerate(it, start=2):
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue

        def g(col):
            return limpiar_texto(fila[H[col]])

        codigo = g("Codigo Cliente")
        if not codigo:
            rep.add("sin_codigo_cliente", f"fila {nro}", "fila descartada")
            continue

        # ---- Cliente (primera aparición gana; el perfilado confirmó que los datos maestros no
        # divergen entre filas del mismo código, así que no hace falta elegir).
        if codigo not in clientes:
            razon = g("Razon Social")
            if not razon:
                razon = f"SIN NOMBRE ({codigo})"
                rep.add("razon_social_vacia", codigo, "se completó con un texto de relleno")

            cond_excel = g("Condicion IVA").upper()
            cond = COND_IVA.get(cond_excel)
            if cond is None:
                cond = "Consumidor Final"
                rep.add("cond_iva_desconocida", codigo, f"{cond_excel!r} -> Consumidor Final")
            vistos_cond[cond] += 1

            # El CUIT solo vale si son 11 dígitos. El padrón trae 1.213 celdas con el relleno
            # '-        -' (guión, espacios, guión), que al quitarle los no-dígitos queda vacío y no
            # se reporta porque nunca fue un CUIT; y 25 filas con 1, 2, 8, 9 o 10 dígitos, que sí se
            # descartan avisando. El documento real ya viaja en numdoc.
            cuit_bruto = re.sub(r"\D", "", g("cuit"))
            if cuit_bruto and len(cuit_bruto) != 11:
                rep.add("cuit_invalido", codigo, f"{g('cuit')!r} ({len(cuit_bruto)} dígitos) -> vacío")
                cuit_bruto = ""

            email = g("email")
            if email and "@" not in email:
                rep.add("email_invalido", codigo, f"{email!r} -> vacío")
                email = ""

            clientes[codigo] = {
                "CodigoInt": recortar(codigo, "CodigoInt", codigo, rep),
                "Cuit": cuit_bruto or None,
                "Documento": recortar(g("numdoc"), "Documento", codigo, rep) or None,
                "Descripcion": recortar(razon, "Descripcion", codigo, rep),
                "CondIvaDescripcion": cond,
                "PermitePresupuesto": g("Admite Presupuesto") == "1",
                "Activo": g("Estado").upper() == "ACTIVO",
                "Domicilio": recortar(g("Domicilio"), "Domicilio", codigo, rep) or None,
                "CodigoPostal": recortar(g("copostal"), "CodigoPostal", codigo, rep) or None,
                "Localidad": recortar(g("localidad"), "Localidad", codigo, rep) or None,
                "Email": recortar(email, "Email", codigo, rep) or None,
            }

        # ---- Tarjeta
        nro_tarjeta = g("Codigo Tarjeta")
        if not nro_tarjeta:
            rep.add("fila_sin_tarjeta", codigo, "el cliente se importa sin tarjeta")
            continue
        tipo_excel = g("tipo_tarjeta").upper()
        tipo = TIPO_TARJETA.get(tipo_excel)
        if tipo is None:
            tipo = TIPO_TARJETA_DEFECTO
            rep.add("tipo_tarjeta_ausente", codigo,
                    f"{tipo_excel!r} en tarjeta {nro_tarjeta} -> {TIPO_TARJETA_DEFECTO}")
        tarjetas.add((codigo, tipo, recortar(nro_tarjeta, "NroTarjeta", codigo, rep)))

    wb.close()

    clientes, tarjetas = resolver_choques_de_caso(clientes, tarjetas, rep)

    # Una misma tarjeta en clientes distintos es un dato imposible para una tarjeta de fidelización:
    # se avisa y se conservan todas (el PK las admite), para que se resuelva a mano.
    dueños = defaultdict(set)
    for cod, _tipo, nro in tarjetas:
        dueños[nro].add(cod)
    for nro, cods in dueños.items():
        if len(cods) > 1:
            rep.add("tarjeta_en_varios_clientes", ",".join(sorted(cods)), f"tarjeta {nro}")

    return clientes, tarjetas, vistos_cond


def resolver_choques_de_caso(clientes, tarjetas, rep):
    """
    El índice único de Clientes.CodigoInt es case-insensitive (collation por defecto de SQL Server),
    pero el ERP anterior distingue códigos por mayúscula/minúscula: en el padrón, 'C9840' y 'c9840'
    son DOS clientes distintos (GUO XIAFENG y MARTINEZ MARIA ALEJANDRA, con documento y tarjeta
    propios). Fusionarlos perdería un cliente y descartarlo dejaría su tarjeta muerta en la caja, así
    que al segundo se le agrega un sufijo y se reporta para que se corrija a mano.

    Se conserva sin tocar el código que ya viene en mayúsculas (orden determinístico), de modo que
    reejecutar la importación produzca siempre el mismo resultado.
    """
    grupos = defaultdict(list)
    for codigo in clientes:
        grupos[codigo.upper()].append(codigo)

    renombres = {}
    for clave, codigos in grupos.items():
        if len(codigos) == 1:
            continue
        # El que ya está en mayúsculas conserva su código; el resto, en orden, recibe sufijo.
        codigos.sort(key=lambda c: (c != c.upper(), c))
        for i, codigo in enumerate(codigos[1:], start=2):
            nuevo = f"{codigo.upper()}-{i}"
            while nuevo.upper() in grupos:
                i += 1
                nuevo = f"{codigo.upper()}-{i}"
            renombres[codigo] = nuevo
            rep.add("choque_de_mayusculas", codigo,
                    f"choca con {codigos[0]!r} en el índice único (que no distingue caso) "
                    f"pero es otro cliente: se importa como {nuevo!r}")

    if not renombres:
        return clientes, tarjetas

    clientes_final = {}
    for codigo, c in clientes.items():
        final = renombres.get(codigo, codigo)
        c["CodigoInt"] = final
        clientes_final[final] = c
    tarjetas_final = {(renombres.get(cod, cod), tipo, nro) for cod, tipo, nro in tarjetas}
    return clientes_final, tarjetas_final


def conectar(conn_str):
    m = dict(re.findall(r"([^=;]+)=([^;]*)", conn_str))
    m = {k.strip().lower(): v.strip() for k, v in m.items()}
    server = m.get("server") or m.get("data source")
    base = m.get("database") or m.get("initial catalog")
    usuario = m.get("user id") or m.get("uid")
    clave = m.get("password") or m.get("pwd")
    odbc = (f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={base};"
            f"UID={usuario};PWD={clave};TrustServerCertificate=yes;Encrypt=yes")
    return pyodbc.connect(odbc, autocommit=False)


def cargar(cn, clientes, tarjetas, rep):
    cur = cn.cursor()
    cur.fast_executemany = True

    # ---- Catálogo de condiciones de IVA: crear las que falten.
    cur.execute("SELECT Descripcion, IdCondIva FROM dbo.CondicionesIva")
    cond_ids = {d.strip(): i for d, i in cur.fetchall()}
    for desc, (letra, codigo) in COND_IVA_NUEVAS.items():
        if desc not in cond_ids:
            cur.execute(
                "INSERT INTO dbo.CondicionesIva (Descripcion, Letra, CodigoInterno, CreatedAtUtc) "
                "VALUES (?, ?, ?, SYSUTCDATETIME())", desc, letra, codigo)
            cur.execute("SELECT IdCondIva FROM dbo.CondicionesIva WHERE Descripcion = ?", desc)
            cond_ids[desc] = cur.fetchone()[0]
            print(f"  catálogo: creada condición de IVA {desc!r} (letra {letra})")

    faltantes = {c["CondIvaDescripcion"] for c in clientes.values()} - set(cond_ids)
    if faltantes:
        raise SystemExit(f"Faltan condiciones de IVA en el catálogo: {faltantes}")

    cur.execute("SELECT Descripcion, IdTipoTarjeta FROM dbo.TiposTarjeta")
    tipo_ids = {d.strip(): i for d, i in cur.fetchall()}
    faltantes = {t for _c, t, _n in tarjetas} - set(tipo_ids)
    if faltantes:
        raise SystemExit(f"Faltan tipos de tarjeta en el catálogo: {faltantes}")

    # ---- Códigos que ya existen: se actualizan en vez de duplicarse (CodigoInt es único).
    # La comparación va en mayúsculas porque así compara el índice único de SQL Server: si acá se
    # comparara distinguiendo caso, un código que ya existe con otra capitalización se intentaría
    # insertar y rompería toda la carga por clave duplicada.
    cur.execute("SELECT CodigoInt, IdCliente FROM dbo.Clientes")
    existentes = {c.strip().upper(): i for c, i in cur.fetchall()}
    a_insertar = {k: v for k, v in clientes.items() if k.upper() not in existentes}
    a_actualizar = {k: v for k, v in clientes.items() if k.upper() in existentes}
    for k in a_actualizar:
        rep.add("codigo_ya_existente", k, "se actualizó el cliente existente")

    cur.execute("SELECT ISNULL(MAX(IdCliente), 0) FROM dbo.Clientes")
    siguiente = cur.fetchone()[0] + 1

    ids = dict(existentes)   # claves en mayúsculas
    filas = []
    for i, (codigo, c) in enumerate(sorted(a_insertar.items())):
        idc = siguiente + i
        ids[codigo.upper()] = idc
        filas.append((idc, c["CodigoInt"], c["Cuit"], c["Documento"], c["Descripcion"],
                      cond_ids[c["CondIvaDescripcion"]], c["PermitePresupuesto"], c["Activo"],
                      c["Domicilio"], c["CodigoPostal"], c["Localidad"], c["Email"]))

    print(f"  insertando {len(filas)} clientes (IdCliente {siguiente}..{siguiente + len(filas) - 1})")
    cur.execute("SET IDENTITY_INSERT dbo.Clientes ON")
    cur.executemany(
        "INSERT INTO dbo.Clientes (IdCliente, CodigoInt, Cuit, Documento, Descripcion, IdCondIva, "
        "PermitePresupuesto, Activo, Domicilio, CodigoPostal, Localidad, Email, CreatedAtUtc, CreatedBy) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?, SYSUTCDATETIME(), 'importacion-padron')", filas)
    cur.execute("SET IDENTITY_INSERT dbo.Clientes OFF")

    if a_actualizar:
        upd = [(c["Cuit"], c["Documento"], c["Descripcion"], cond_ids[c["CondIvaDescripcion"]],
                c["PermitePresupuesto"], c["Activo"], c["Domicilio"], c["CodigoPostal"],
                c["Localidad"], c["Email"], existentes[k.upper()]) for k, c in a_actualizar.items()]
        print(f"  actualizando {len(upd)} clientes que ya existían")
        cur.executemany(
            "UPDATE dbo.Clientes SET Cuit=?, Documento=?, Descripcion=?, IdCondIva=?, "
            "PermitePresupuesto=?, Activo=?, Domicilio=?, CodigoPostal=?, Localidad=?, Email=?, "
            "UpdatedAtUtc=SYSUTCDATETIME(), UpdatedBy='importacion-padron' WHERE IdCliente=?", upd)

    # ---- Tarjetas: se saltean las que ya estuvieran cargadas (la PK es idCliente+tipo+nro).
    cur.execute("SELECT IdCliente, IdTipoTarjeta, NroTarjeta FROM dbo.TarjetasClientes")
    ya = {(a, b, c.strip()) for a, b, c in cur.fetchall()}
    filas_t = []
    for codigo, tipo, nro in sorted(tarjetas):
        clave = (ids[codigo.upper()], tipo_ids[tipo], nro)
        if clave not in ya:
            ya.add(clave)
            filas_t.append(clave)

    print(f"  insertando {len(filas_t)} tarjetas")
    cur.executemany(
        "INSERT INTO dbo.TarjetasClientes (IdCliente, IdTipoTarjeta, NroTarjeta, CreatedAtUtc, CreatedBy) "
        "VALUES (?,?,?, SYSUTCDATETIME(), 'importacion-padron')", filas_t)

    return len(filas), len(a_actualizar), len(filas_t)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True)
    p.add_argument("--conn", default=os.environ.get("POS_CONN"))
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--reporte", default="anomalias-padron.csv")
    args = p.parse_args()

    if not args.dry_run and not args.conn:
        sys.exit("Falta la cadena de conexión: usá --conn o la variable POS_CONN.")

    rep = Reporte()
    print(f"Leyendo {args.excel} ...")
    clientes, tarjetas, cond = transformar(args.excel, rep)
    print(f"  {len(clientes)} clientes distintos, {len(tarjetas)} tarjetas distintas")
    print("  condición de IVA:", dict(cond.most_common()))
    print("  tipo de tarjeta:", dict(Counter(t for _c, t, _n in tarjetas).most_common()))

    print("\nAnomalías:")
    for tipo, n in rep.resumen():
        print(f"  {tipo:32} {n}")
    rep.volcar(args.reporte)
    print(f"  detalle en {args.reporte}")

    if args.dry_run:
        print("\n--dry-run: no se tocó la base.")
        return

    cn = conectar(args.conn)
    try:
        print("\nCargando ...")
        ins, upd, tar = cargar(cn, clientes, tarjetas, rep)
        cn.commit()
        print(f"\nListo: {ins} clientes nuevos, {upd} actualizados, {tar} tarjetas.")
    except Exception:
        cn.rollback()
        print("\nERROR: se revirtió toda la carga.")
        raise
    finally:
        cn.close()


if __name__ == "__main__":
    main()
